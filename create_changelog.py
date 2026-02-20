from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()
ws = wb.active
ws.title = "Changelog"

# Headers
headers = ["Generated Lines in File", "User Name", "Date and Time (UTC)"]
ws.append(headers)

# Style headers
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)

for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Add data
ws.append([30, "pramilavhalyal", "2026-02-20 07:17:00"])

# Format columns
ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 25

# Save
wb.save("changelog_extract.xlsx")
print("✓ File created: changelog_extract.xlsx")